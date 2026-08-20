"""
Excel Logger

Creates:
    logs/rag_logs.xlsx

One row = One chatbot request.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Any

from openpyxl import Workbook, load_workbook

from config import EXCEL_LOG_FILE

logger = logging.getLogger(__name__)


class ExcelLogger:

    SHEET_NAME = "RAG_Logs"

    HEADERS = [

        "Timestamp",

        "Status",

        "Question",

        "Answer",

        "Confidence",

        "Retrieval Time (ms)",

        "LLM Time (ms)",

        "Total Time (ms)",

        "Input Tokens",

        "Output Tokens",

        "Total Tokens",

        "Input Cost",

        "Output Cost",

        "Total Cost",

        "Sources",

        "Error"

    ]

    def __init__(self):

        self.file_path = Path(EXCEL_LOG_FILE)

        self._initialize_workbook()

    # ==========================================================
    # Workbook Initialization
    # ==========================================================

    def _initialize_workbook(self):

        if self.file_path.exists():
            return

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = self.SHEET_NAME

        sheet.append(self.HEADERS)

        workbook.save(self.file_path)

        logger.info(
            "Excel log created: %s",
            self.file_path,
        )

    # ==========================================================
    # Log Success
    # ==========================================================

    def log(
        self,
        payload: Dict[str, Any],
    ) -> None:

        workbook = load_workbook(self.file_path)

        sheet = workbook[self.SHEET_NAME]

        sources = ""

        if payload.get("sources"):

            sources = ", ".join(

                str(src.get("row_number", ""))

                for src in payload["sources"]

            )

        row = [

            payload.get("logged_at"),

            payload.get("status"),

            payload.get("question"),

            payload.get("answer"),

            payload.get("confidence"),

            payload.get("retrieval_time_ms"),

            payload.get("llm_time_ms"),

            payload.get("total_time_ms"),

            payload.get("input_tokens"),

            payload.get("output_tokens"),

            payload.get("total_tokens"),

            payload.get("cost_breakdown", {}).get("input_cost"),

            payload.get("cost_breakdown", {}).get("output_cost"),

            payload.get("cost"),

            sources,

            ""

        ]

        sheet.append(row)

        workbook.save(self.file_path)

    # ==========================================================
    # Log Failure
    # ==========================================================

    def log_error(
        self,
        payload: Dict[str, Any],
    ) -> None:

        workbook = load_workbook(self.file_path)

        sheet = workbook[self.SHEET_NAME]

        row = [

            payload.get("logged_at"),

            payload.get("status"),

            payload.get("question"),

            "",

            "",

            "",

            "",

            "",

            "",

            "",

            "",

            "",

            "",

            "",

            "",

            payload.get("error")

        ]

        sheet.append(row)

        workbook.save(self.file_path)